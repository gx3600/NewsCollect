"""Configuration management — loads sources.yaml and settings.yaml."""

import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import yaml

logger = logging.getLogger(__name__)


@dataclass
class SourceConfig:
    """Configuration for a single news source."""

    name: str
    url: str
    enabled: bool = True
    interval: int = 300              # seconds between crawl runs
    use_rss: bool = False            # RSS/Atom feed source (not HTML scraping)
    use_stealth: bool = False        # use StealthyFetcher
    use_dynamic: bool = False        # use DynamicFetcher (full browser)
    download_delay: float = 1.0      # seconds between requests
    fetch_content: bool = True       # fetch full article body from detail pages
    max_items: int = 10              # max articles per crawl
    selectors: dict = field(default_factory=dict)
    extra: dict = field(default_factory=dict)

    @classmethod
    def from_dict(cls, name: str, data: dict) -> "SourceConfig":
        return cls(
            name=name,
            url=data.get("url", ""),
            enabled=data.get("enabled", True),
            interval=data.get("interval", 300),
            use_rss=data.get("use_rss", False),
            use_stealth=data.get("use_stealth", False),
            use_dynamic=data.get("use_dynamic", False),
            download_delay=data.get("download_delay", 1.0),
            fetch_content=data.get("fetch_content", True),
            max_items=data.get("max_items", 10),
            selectors=data.get("selectors", {}),
            extra=data.get("extra", {}),
        )


class Config:
    """Singleton configuration loader.

    Usage:
        cfg = Config()
        source_cfg = cfg.get_source("eastmoney")
        all_sources = cfg.sources
    """

    _instance: Optional["Config"] = None

    def __new__(cls) -> "Config":
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._instance._loaded = False
        return cls._instance

    def __init__(self):
        if self._loaded:
            return
        self._loaded = True
        self._load()

    def _load(self):
        base = Path(__file__).parent.parent.parent  # NewsCollect/

        # Load global settings
        settings_path = base / "config" / "settings.yaml"
        if settings_path.exists():
            with open(settings_path, "r", encoding="utf-8") as f:
                self.settings = yaml.safe_load(f) or {}
        else:
            self.settings = {}

        # Load source configurations
        sources_path = base / "config" / "sources.yaml"
        if sources_path.exists():
            with open(sources_path, "r", encoding="utf-8") as f:
                raw = yaml.safe_load(f) or {}
        else:
            raw = {}

        self._sources: dict[str, SourceConfig] = {}
        for name, data in raw.get("sources", {}).items():
            self._sources[name] = SourceConfig.from_dict(name, data)

    # ── accessors ──────────────────────────────────────────

    @property
    def sources(self) -> dict[str, SourceConfig]:
        return self._sources

    @property
    def enabled_sources(self) -> dict[str, SourceConfig]:
        return {k: v for k, v in self._sources.items() if v.enabled}

    def get_source(self, name: str) -> Optional[SourceConfig]:
        return self._sources.get(name)

    # ── global settings helpers ────────────────────────────

    @property
    def default_concurrency(self) -> int:
        return self.settings.get("concurrency", 5)

    @property
    def default_timeout(self) -> int:
        return self.settings.get("timeout", 30)

    @property
    def default_user_agent(self) -> str:
        return self.settings.get(
            "user_agent",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        )

    @property
    def proxy_list(self) -> list[str]:
        return self.settings.get("proxies", [])

    @property
    def db_path(self) -> str:
        return self.settings.get("db_path", "data/news.db")

    @property
    def retention_days(self) -> int:
        return self.settings.get("retention_days", 90)

    @property
    def log_level(self) -> str:
        return self.settings.get("log_level", "INFO")

    # ── DeepSeek API settings ──────────────────────────────

    @property
    def deepseek_api_key(self) -> str:
        return self.settings.get("deepseek", {}).get("api_key", "")

    @property
    def deepseek_base_url(self) -> str:
        return self.settings.get("deepseek", {}).get(
            "base_url", "https://api.deepseek.com/v1"
        )

    @property
    def deepseek_model(self) -> str:
        return self.settings.get("deepseek", {}).get("model", "deepseek-chat")

    @property
    def deepseek_max_tokens(self) -> int:
        return self.settings.get("deepseek", {}).get("max_tokens", 4096)

    @property
    def deepseek_temperature(self) -> float:
        return self.settings.get("deepseek", {}).get("temperature", 0.1)

    @property
    def deepseek_timeout(self) -> int:
        return self.settings.get("deepseek", {}).get("timeout", 120)

    @property
    def deepseek_max_retries(self) -> int:
        return self.settings.get("deepseek", {}).get("max_retries", 3)

    # ── analysis settings ──────────────────────────────────

    @property
    def analysis_interval(self) -> int:
        return self.settings.get("analysis", {}).get("interval", 30)

    @property
    def analysis_batch_size(self) -> int:
        return self.settings.get("analysis", {}).get("batch_size", 50)

    @property
    def analysis_concurrency(self) -> int:
        return self.settings.get("analysis", {}).get("concurrency", 50)

    # ── futures variety whitelist ──────────────────────────

    @property
    def futures_variety_path(self) -> Path:
        """Path to the futures variety Excel file."""
        configured = self.settings.get("futures_variety_path", "config/futures_variety.xlsx")
        p = Path(configured)
        if not p.is_absolute():
            base = Path(__file__).parent.parent.parent  # NewsCollect/
            p = base / p
        return p

    @property
    def futures_varieties(self) -> list[str]:
        """Deduplicated list of Chinese futures contract names from the Excel file."""
        varieties = self._load_variety_data()
        return varieties["names"]

    @property
    def futures_varieties_str(self) -> str:
        """Comma-separated Chinese futures variety names, for prompt injection."""
        return "、".join(self.futures_varieties)

    @property
    def futures_variety_keywords_map(self) -> dict[str, list[str]]:
        """Mapping: {contractName: [keyword list]} from the Excel."""
        varieties = self._load_variety_data()
        return varieties["keywords"]

    # cache for _load_variety_data
    _variety_cache: Optional[dict] = None

    def _load_variety_data(self) -> dict:
        """Load and parse the futures variety Excel file. Results are cached."""
        if self._variety_cache is not None:
            return self._variety_cache

        names: list[str] = []
        keywords_map: dict[str, list[str]] = {}

        try:
            import openpyxl
        except ImportError:
            logger.warning("openpyxl not installed — futures variety list unavailable.")
            self._variety_cache = {"names": names, "keywords": keywords_map}
            return self._variety_cache

        xlsx_path = self.futures_variety_path
        if not xlsx_path.exists():
            logger.warning(
                f"Futures variety file not found: {xlsx_path} — "
                f"variety whitelist will be empty."
            )
            self._variety_cache = {"names": names, "keywords": keywords_map}
            return self._variety_cache

        try:
            wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]

            # Locate columns
            col_idx: dict[str, int] = {}
            for i, h in enumerate(headers):
                if h == "contractName":
                    col_idx["name"] = i
                elif h == "keyWords":
                    col_idx["keywords"] = i

            if "name" not in col_idx:
                logger.error(f"contractName column not found in {xlsx_path}")
                wb.close()
                self._variety_cache = {"names": names, "keywords": keywords_map}
                return self._variety_cache

            seen_names: set[str] = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                name = str(row[col_idx["name"]]).strip() if row[col_idx["name"]] else ""
                if not name or name in seen_names:
                    continue
                seen_names.add(name)
                names.append(name)

                # Parse keywords
                if "keywords" in col_idx:
                    kw_raw = row[col_idx["keywords"]]
                    if kw_raw:
                        kws = [k.strip() for k in str(kw_raw).split(",") if k.strip()]
                        keywords_map[name] = kws
                    else:
                        keywords_map[name] = []

            wb.close()
            logger.info(f"Loaded {len(names)} futures varieties from {xlsx_path}")

        except Exception as e:
            logger.error(f"Failed to load futures varieties from {xlsx_path}: {e}")

        self._variety_cache = {"names": names, "keywords": keywords_map}
        return self._variety_cache
